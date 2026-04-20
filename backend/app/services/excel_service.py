from __future__ import annotations

from datetime import datetime
import json
import os
from pathlib import Path
import re
import shutil
from tempfile import NamedTemporaryFile
import threading

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

try:
    import olefile
except ImportError:  # pragma: no cover
    olefile = None

ROOT_DIR = Path(__file__).resolve().parents[3]
DATA_DIR = ROOT_DIR / 'data'
ORACLE_PATH = DATA_DIR / 'Oracle.xlsx'
SHE_PATH = DATA_DIR / 'SHE.xlsx'
NOTES_PATH = DATA_DIR / 'Notes.xlsx'

DEFAULT_ADMIN_EMAILS = 'msp@haycarb.com'
NOTE_COLUMNS = [
    'id',
    'timestamp',
    'userEmail',
    'nic',
    'message',
    'status',
    'adminEmail',
    'adminComment',
    'requestType',
    'targetRecordId',
    'payloadJson',
]

EMAIL_PATTERN = re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]+$')
NIC_PATTERN = re.compile(r'^[A-Za-z0-9-]{8,20}$')

_sheet_lock = threading.Lock()


def _safe(value: object) -> str:
    if value is None:
        return ''
    return str(value).strip()


def _is_email(value: object) -> bool:
    return bool(EMAIL_PATTERN.match(_safe(value).lower()))


def _is_nic(value: object) -> bool:
    candidate = _safe(value)
    return bool(candidate) and bool(NIC_PATTERN.match(candidate)) and len(candidate) >= 10


def _to_number(value: object) -> float:
    if value in (None, ''):
        return 0.0

    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _admin_email_set() -> set[str]:
    raw = os.getenv('ADMIN_EMAILS', DEFAULT_ADMIN_EMAILS)
    return {entry.strip().lower() for entry in raw.split(',') if entry.strip()}


def is_admin_email(email: str) -> bool:
    return email.strip().lower() in _admin_email_set()


def _ensure_notes_file() -> None:
    if NOTES_PATH.exists():
        return

    wb = Workbook()
    ws = wb.active
    ws.title = 'Notes'
    ws.append(NOTE_COLUMNS)
    wb.save(NOTES_PATH)


def _normalize_date(value: object) -> str:
    if value is None or _safe(value) == '':
        return ''

    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')

    text = _safe(value)
    for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d', '%m-%d-%Y'):
        try:
            return datetime.strptime(text, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue

    return text


def _normalize_gender(value: object) -> str:
    text = _safe(value).lower()
    if text in ('m', 'male'):
        return 'Male'
    if text in ('f', 'female'):
        return 'Female'
    return 'Other'


def _ensure_files() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    if not ORACLE_PATH.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = 'Oracle'
        ws.append(['Email', 'NIC'])
        wb.save(ORACLE_PATH)

    if not SHE_PATH.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = 'SHE'
        ws.append(
            [
                'id',
                'name',
                'nic',
                'dob',
                'gender',
                'relation',
                'category',
                'effectiveDate',
                'grade',
                'totalPremium',
                'employeeVoluntaryEnhanceLimit',
                'note',
            ]
        )
        wb.save(SHE_PATH)

    _ensure_notes_file()


def _detect_she_layout(ws) -> str:
    header_values: list[str] = []
    for row in ws.iter_rows(min_row=1, max_row=4, values_only=True):
        header_values.extend([_safe(cell).lower() for cell in row if _safe(cell)])

    if 'id' in header_values and 'nic' in header_values:
        return 'seeded'

    if any('nic number' in value for value in header_values):
        return 'real'

    return 'real'


def _build_record_from_seeded(row_idx: int, row: tuple[object, ...]) -> dict | None:
    record_id = _safe(row[0]) if len(row) > 0 else ''
    nic = _safe(row[2]) if len(row) > 2 else ''
    if not record_id or not _is_nic(nic):
        return None

    return {
        'id': record_id,
        'name': _safe(row[1]) if len(row) > 1 else '',
        'nic': nic,
        'dob': _normalize_date(row[3]) if len(row) > 3 else '',
        'gender': _normalize_gender(row[4]) if len(row) > 4 else 'Other',
        'relation': _safe(row[5]) if len(row) > 5 else '',
        'category': _safe(row[6]) if len(row) > 6 else '',
        'effectiveDate': _normalize_date(row[7]) if len(row) > 7 else '',
        'grade': _safe(row[8]) if len(row) > 8 else '',
        'totalPremium': _to_number(row[9]) if len(row) > 9 else 0.0,
        'employeeVoluntaryEnhanceLimit': _to_number(row[10]) if len(row) > 11 else 0.0,
        'note': _safe(row[11]) if len(row) > 11 else (_safe(row[10]) if len(row) > 10 else ''),
    }


def _build_record_from_real(row_idx: int, row: tuple[object, ...]) -> dict | None:
    nic = _safe(row[0]) if len(row) > 0 else ''
    if not _is_nic(nic):
        return None

    return {
        'id': f'she-{row_idx}',
        'name': _safe(row[1]) if len(row) > 1 else '',
        'nic': nic,
        'dob': _normalize_date(row[2]) if len(row) > 2 else '',
        'gender': _normalize_gender(row[3]) if len(row) > 3 else 'Other',
        'relation': _safe(row[4]) if len(row) > 4 else '',
        'category': _safe(row[5]) if len(row) > 5 else '',
        'effectiveDate': _normalize_date(row[6]) if len(row) > 6 else '',
        'grade': _safe(row[7]) if len(row) > 7 else '',
        'totalPremium': _to_number(row[12]) if len(row) > 12 else 0.0,
        'employeeVoluntaryEnhanceLimit': _to_number(row[18]) if len(row) > 18 else 0.0,
        'note': _safe(row[26]) if len(row) > 26 else '',
    }


def get_nic_by_email(email: str) -> str | None:
    _ensure_files()
    wb = load_workbook(ORACLE_PATH, read_only=True, data_only=True)
    ws = wb.active
    target = email.lower().strip()

    for row in ws.iter_rows(values_only=True):
        row_values = [_safe(value) for value in row]
        lowered = [value.lower() for value in row_values]
        if target not in lowered:
            continue

        nic_candidates = [value for value in row_values if _is_nic(value)]
        wb.close()
        if nic_candidates:
            return nic_candidates[-1]

    wb.close()
    return None


def get_records_by_nic(nic: str) -> list[dict]:
    _ensure_files()
    wb = load_workbook(SHE_PATH, read_only=True, data_only=True)
    ws = wb.active
    layout = _detect_she_layout(ws)

    records: list[dict] = []
    target_nic = nic.strip().lower()
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if layout == 'seeded':
            record = _build_record_from_seeded(row_idx, row)
        else:
            record = _build_record_from_real(row_idx, row)

        if record and record['nic'].lower() == target_nic:
            records.append(record)

    wb.close()
    return records


def split_employee_dependants(records: list[dict]) -> tuple[dict | None, list[dict]]:
    if not records:
        return None, []

    employee = next(
        (
            record
            for record in records
            if 'employee' in _safe(record.get('relation', '')).lower()
        ),
        records[0],
    )
    dependants = [record for record in records if record['id'] != employee['id']]
    return employee, dependants


def create_dependant(payload: dict) -> str:
    with _sheet_lock:
        _ensure_files()
        wb = load_workbook(SHE_PATH)
        ws = wb.active
        layout = _detect_she_layout(ws)
        new_row_index = ws.max_row + 1

        if layout == 'seeded':
            ws.cell(row=new_row_index, column=1, value=f'row-{new_row_index - 1}')
            ws.cell(row=new_row_index, column=2, value=payload['name'])
            ws.cell(row=new_row_index, column=3, value=payload['nic'])
            ws.cell(row=new_row_index, column=4, value=payload['dob'])
            ws.cell(row=new_row_index, column=5, value=payload['gender'])
            ws.cell(row=new_row_index, column=6, value=payload['relation'])
            ws.cell(row=new_row_index, column=7, value=payload.get('category', 'Dependant'))
            ws.cell(row=new_row_index, column=8, value=payload.get('effectiveDate', ''))
            ws.cell(row=new_row_index, column=9, value=payload.get('grade', ''))
            ws.cell(row=new_row_index, column=10, value=payload.get('totalPremium', 0))
            ws.cell(row=new_row_index, column=11, value=payload.get('employeeVoluntaryEnhanceLimit', 0))
            ws.cell(row=new_row_index, column=12, value=payload.get('note', ''))
            record_id = f'row-{new_row_index - 1}'
        else:
            ws.cell(row=new_row_index, column=1, value=payload['nic'])
            ws.cell(row=new_row_index, column=2, value=payload['name'])
            ws.cell(row=new_row_index, column=3, value=payload['dob'])
            ws.cell(row=new_row_index, column=4, value=payload['gender'])
            ws.cell(row=new_row_index, column=5, value=payload['relation'])
            ws.cell(row=new_row_index, column=6, value=payload.get('category', 'Dependant'))
            ws.cell(row=new_row_index, column=7, value=payload.get('effectiveDate', ''))
            ws.cell(row=new_row_index, column=8, value=payload.get('grade', ''))
            ws.cell(row=new_row_index, column=13, value=payload.get('totalPremium', 0))
            ws.cell(row=new_row_index, column=19, value=payload.get('employeeVoluntaryEnhanceLimit', 0))
            ws.cell(row=new_row_index, column=27, value=payload.get('note', ''))
            record_id = f'she-{new_row_index}'

        wb.save(SHE_PATH)
        wb.close()
        return record_id


def update_record(record_id: str, payload: dict) -> bool:
    with _sheet_lock:
        _ensure_files()
        wb = load_workbook(SHE_PATH)
        ws = wb.active
        layout = _detect_she_layout(ws)

        row_index = -1
        if layout == 'seeded':
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
                if _safe(row[0]) == record_id:
                    row_index = row[0].row if hasattr(row[0], 'row') else -1
                    break

            if row_index == -1:
                for idx in range(1, ws.max_row + 1):
                    if _safe(ws.cell(row=idx, column=1).value) == record_id:
                        row_index = idx
                        break
        elif record_id.startswith('she-'):
            try:
                row_index = int(record_id.split('-', 1)[1])
            except ValueError:
                row_index = -1

        if row_index < 1 or row_index > ws.max_row:
            wb.close()
            return False

        seeded_map = {
            'name': 2,
            'nic': 3,
            'dob': 4,
            'gender': 5,
            'relation': 6,
            'category': 7,
            'effectiveDate': 8,
            'grade': 9,
            'totalPremium': 10,
            'employeeVoluntaryEnhanceLimit': 11,
            'note': 12,
        }
        real_map = {
            'nic': 1,
            'name': 2,
            'dob': 3,
            'gender': 4,
            'relation': 5,
            'category': 6,
            'effectiveDate': 7,
            'grade': 8,
            'totalPremium': 13,
            'employeeVoluntaryEnhanceLimit': 19,
            'note': 27,
        }

        key_map = seeded_map if layout == 'seeded' else real_map
        for key, value in payload.items():
            column = key_map.get(key)
            if column and value is not None:
                ws.cell(row=row_index, column=column, value=value)

        wb.save(SHE_PATH)
        wb.close()
        return True


def submit_user_note(user_email: str, nic: str, message: str) -> str:
    with _sheet_lock:
        _ensure_files()
        wb = load_workbook(NOTES_PATH)
        ws = wb.active

        note_id = f'note-{ws.max_row}'
        ws.append(
            [
                note_id,
                datetime.now().isoformat(timespec='seconds'),
                user_email.strip(),
                nic.strip(),
                message.strip(),
                'NEW',
                '',
                '',
                'NOTE',
                '',
                '',
            ]
        )

        wb.save(NOTES_PATH)
        wb.close()
        return note_id


def submit_change_request(user_email: str, nic: str, target_record_id: str, changes: dict) -> str:
    with _sheet_lock:
        _ensure_files()
        wb = load_workbook(NOTES_PATH)
        ws = wb.active

        note_id = f'note-{ws.max_row}'
        message = 'Profile change request submitted by user.'
        ws.append(
            [
                note_id,
                datetime.now().isoformat(timespec='seconds'),
                user_email.strip(),
                nic.strip(),
                message,
                'NEW',
                '',
                '',
                'CHANGE',
                target_record_id.strip(),
                json.dumps(changes),
            ]
        )

        wb.save(NOTES_PATH)
        wb.close()
        return note_id


def submit_dependant_notification(
    user_email: str,
    nic: str,
    message: str,
    target_record_id: str,
    payload: dict,
    request_type: str,
) -> str:
    with _sheet_lock:
        _ensure_files()
        wb = load_workbook(NOTES_PATH)
        ws = wb.active

        note_id = f'note-{ws.max_row}'
        ws.append(
            [
                note_id,
                datetime.now().isoformat(timespec='seconds'),
                user_email.strip(),
                nic.strip(),
                message.strip(),
                'NEW',
                '',
                '',
                request_type.strip().upper(),
                target_record_id.strip(),
                json.dumps(payload),
            ]
        )

        wb.save(NOTES_PATH)
        wb.close()
        return note_id


def list_notes(status: str | None = None) -> list[dict]:
    _ensure_files()
    wb = load_workbook(NOTES_PATH, read_only=True, data_only=True)
    ws = wb.active

    results: list[dict] = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx == 1:
            continue

        note = {
            'id': _safe(row[0]) if len(row) > 0 else '',
            'timestamp': _safe(row[1]) if len(row) > 1 else '',
            'userEmail': _safe(row[2]) if len(row) > 2 else '',
            'nic': _safe(row[3]) if len(row) > 3 else '',
            'message': _safe(row[4]) if len(row) > 4 else '',
            'status': _safe(row[5]) if len(row) > 5 else 'NEW',
            'adminEmail': _safe(row[6]) if len(row) > 6 else '',
            'adminComment': _safe(row[7]) if len(row) > 7 else '',
            'requestType': _safe(row[8]) if len(row) > 8 else 'NOTE',
            'targetRecordId': _safe(row[9]) if len(row) > 9 else '',
            'payloadJson': _safe(row[10]) if len(row) > 10 else '',
        }

        if not note['id']:
            continue
        if status and status.upper() != 'ALL' and note['status'].upper() != status.upper():
            continue
        results.append(note)

    wb.close()
    results.sort(key=lambda entry: entry['timestamp'], reverse=True)
    return results


def delete_note(note_id: str) -> bool:
    with _sheet_lock:
        _ensure_files()
        wb = load_workbook(NOTES_PATH)
        ws = wb.active

        delete_row = -1
        for idx in range(2, ws.max_row + 1):
            if _safe(ws.cell(row=idx, column=1).value) == note_id:
                delete_row = idx
                break

        if delete_row == -1:
            wb.close()
            return False

        ws.delete_rows(delete_row, 1)
        wb.save(NOTES_PATH)
        wb.close()
        return True


def clear_notes(status: str = 'RESOLVED') -> int:
    with _sheet_lock:
        _ensure_files()
        wb = load_workbook(NOTES_PATH)
        ws = wb.active

        target = status.upper()
        rows_to_delete: list[int] = []
        for idx in range(2, ws.max_row + 1):
            note_status = _safe(ws.cell(row=idx, column=6).value).upper() or 'NEW'
            if target == 'ALL' or note_status == target:
                rows_to_delete.append(idx)

        for idx in reversed(rows_to_delete):
            ws.delete_rows(idx, 1)

        wb.save(NOTES_PATH)
        wb.close()
        return len(rows_to_delete)


def _append_note_to_employee_record(nic: str, message: str, admin_comment: str = '') -> None:
    wb = load_workbook(SHE_PATH)
    ws = wb.active
    layout = _detect_she_layout(ws)

    relation_column = 6 if layout == 'seeded' else 5
    nic_column = 3 if layout == 'seeded' else 1
    note_column = 11 if layout == 'seeded' else 27

    target_row = None
    for row_idx in range(1, ws.max_row + 1):
        row_nic = _safe(ws.cell(row=row_idx, column=nic_column).value)
        relation = _safe(ws.cell(row=row_idx, column=relation_column).value).lower()
        if row_nic.lower() != nic.lower():
            continue
        if 'employee' in relation:
            target_row = row_idx
            break

    if target_row is None:
        wb.close()
        return

    existing = _safe(ws.cell(row=target_row, column=note_column).value)
    admin_suffix = f' | Admin: {admin_comment.strip()}' if admin_comment.strip() else ''
    entry = f"[{datetime.now().strftime('%Y-%m-%d %H:%M')}] {message.strip()}{admin_suffix}".strip()
    updated = f'{existing}\n{entry}'.strip() if existing else entry
    ws.cell(row=target_row, column=note_column, value=updated)

    wb.save(SHE_PATH)
    wb.close()


def resolve_note(
    note_id: str,
    admin_email: str,
    status: str,
    admin_comment: str = '',
    apply_to_sheet: bool = False,
) -> bool:
    apply_payload: dict | None = None
    apply_note: tuple[str, str, str] | None = None

    with _sheet_lock:
        _ensure_files()
        wb = load_workbook(NOTES_PATH)
        ws = wb.active

        row_index = -1
        for idx in range(2, ws.max_row + 1):
            if _safe(ws.cell(row=idx, column=1).value) == note_id:
                row_index = idx
                break

        if row_index == -1:
            wb.close()
            return False

        ws.cell(row=row_index, column=6, value=status.upper())
        ws.cell(row=row_index, column=7, value=admin_email.strip())
        ws.cell(row=row_index, column=8, value=admin_comment.strip())

        nic = _safe(ws.cell(row=row_index, column=4).value)
        message = _safe(ws.cell(row=row_index, column=5).value)
        request_type = _safe(ws.cell(row=row_index, column=9).value).upper() or 'NOTE'
        target_record_id = _safe(ws.cell(row=row_index, column=10).value)
        payload_json = _safe(ws.cell(row=row_index, column=11).value)

        wb.save(NOTES_PATH)
        wb.close()

        if apply_to_sheet:
            if request_type == 'CHANGE' and target_record_id and payload_json:
                try:
                    payload = json.loads(payload_json)
                    if isinstance(payload, dict):
                        apply_payload = {'record_id': target_record_id, 'changes': payload}
                except json.JSONDecodeError:
                    apply_note = (nic, message, admin_comment)
            elif nic and message:
                apply_note = (nic, message, admin_comment)

    if apply_payload:
        update_record(apply_payload['record_id'], apply_payload['changes'])
    elif apply_note:
        _append_note_to_employee_record(apply_note[0], apply_note[1], apply_note[2])

    return True


def get_sheet_preview(limit: int = 200) -> dict:
    _ensure_files()
    wb = load_workbook(SHE_PATH, read_only=True, data_only=True)
    ws = wb.active

    # Some SHE files include banner rows before column headers.
    header_row_index = 1
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
        if any(_safe(value) for value in row):
            header_row_index = idx
            break

    row_stream = ws.iter_rows(min_row=header_row_index, values_only=True)
    header_row = next(row_stream, None)
    if header_row is None:
        wb.close()
        return {'headers': [], 'rows': []}

    detected_columns = max(
        (col_idx for col_idx, cell in enumerate(header_row, start=1) if _safe(cell)),
        default=0,
    )
    column_count = min(max(detected_columns, 1), 120)

    headers: list[str] = []
    for idx in range(1, column_count + 1):
        label = _safe(header_row[idx - 1]) if idx - 1 < len(header_row) else ''
        headers.append(label if label else f'Column {idx}')

    data_rows: list[list[str]] = []
    for row in row_stream:
        rendered = [_safe(row[idx]) if idx < len(row) else '' for idx in range(column_count)]
        if any(rendered):
            data_rows.append(rendered)
        if len(data_rows) >= limit:
            break

    wb.close()
    return {'headers': headers, 'rows': data_rows}


def get_she_file_path() -> Path:
    _ensure_files()
    return SHE_PATH


def get_oracle_file_path() -> Path:
    _ensure_files()
    return ORACLE_PATH


def _validate_uploaded_workbook(uploaded_file_path: Path) -> None:
    last_error: Exception | None = None

    # Some enterprise-generated workbooks fail in read_only mode but load
    # correctly in normal mode. Accept either mode as valid.
    for read_only in (True, False):
        try:
            wb = load_workbook(uploaded_file_path, read_only=read_only, data_only=True)
            if not wb.sheetnames:
                wb.close()
                raise ValueError('Workbook has no sheets.')
            wb.close()
            return
        except Exception as exc:  # pragma: no cover - best-effort compatibility
            last_error = exc

    detail = str(last_error) if last_error else 'Unknown workbook parsing error.'
    raise ValueError(
        f'Workbook cannot be parsed by the server ({detail}). '
        'Please ensure the file is a standard .xlsx and not password-protected.'
    )


def _normalize_uploaded_workbook(uploaded_file_path: Path) -> Path:
    try:
        _validate_uploaded_workbook(uploaded_file_path)
        return uploaded_file_path
    except ValueError as exc:
        parse_error = str(exc)

    # Office-encrypted workbooks expose an OLE container with EncryptedPackage.
    if olefile and olefile.isOleFile(str(uploaded_file_path)):
        try:
            ole = olefile.OleFileIO(str(uploaded_file_path))
            stream_names = {'/'.join(entry) for entry in ole.listdir()}
            ole.close()
            if 'EncryptedPackage' in stream_names:
                raise ValueError(
                    'Workbook is encrypted/protected (IRM or password-protected). '
                    'Please open it in Excel, remove protection, and Save As a normal .xlsx file before uploading.'
                )
        except OSError:
            pass

    # Fallback for legacy/non-zip excel sources by converting to plain .xlsx values.
    try:
        with NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            normalized_path = Path(tmp.name)

        sheets = pd.read_excel(uploaded_file_path, sheet_name=None, header=None, dtype=object)
        wb = Workbook()
        default_ws = wb.active
        wb.remove(default_ws)

        for index, (sheet_name, frame) in enumerate(sheets.items(), start=1):
            ws = wb.create_sheet(title=(sheet_name or f'Sheet{index}')[:31])
            safe_frame = frame.where(pd.notna(frame), None)
            for row in dataframe_to_rows(safe_frame, index=False, header=False):
                ws.append(row)

        if not wb.sheetnames:
            wb.create_sheet(title='Sheet1')

        wb.save(normalized_path)
        wb.close()
        _validate_uploaded_workbook(normalized_path)
        return normalized_path
    except Exception as fallback_error:
        raise ValueError(
            'Workbook cannot be parsed by the server '
            f'({parse_error}). Legacy conversion also failed ({fallback_error}).'
        ) from fallback_error


def replace_she_file(uploaded_file_path: Path) -> None:
    with _sheet_lock:
        _ensure_files()

        normalized_path = _normalize_uploaded_workbook(uploaded_file_path)

        backup_name = f"SHE.backup.{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        backup_path = DATA_DIR / backup_name
        shutil.copy2(SHE_PATH, backup_path)
        shutil.copy2(normalized_path, SHE_PATH)

        if normalized_path != uploaded_file_path:
            normalized_path.unlink(missing_ok=True)


def replace_oracle_file(uploaded_file_path: Path) -> None:
    with _sheet_lock:
        _ensure_files()

        normalized_path = _normalize_uploaded_workbook(uploaded_file_path)

        backup_name = f"Oracle.backup.{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        backup_path = DATA_DIR / backup_name
        shutil.copy2(ORACLE_PATH, backup_path)
        shutil.copy2(normalized_path, ORACLE_PATH)

        if normalized_path != uploaded_file_path:
            normalized_path.unlink(missing_ok=True)
